"""
生成测试用Excel文件

Author: Cost-RAG Team
Created: 2025-11-06
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def create_test_excel(output_path: str, num_projects: int = 3):
    """创建测试用Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "多项目成本对比"
    
    projects = [
        {"name": "金地商业广场", "area": 89727.0, "floors": "地上10层，地下2层"},
        {"name": "芷阳广场", "area": 28000.0, "floors": "地上8层，地下1层"},
        {"name": "西安凯德广场", "area": 45600.0, "floors": "地上12层，地下3层"},
    ]
    
    projects = projects[:min(num_projects, len(projects))]
    
    # 设置列宽
    ws.column_dimensions['A'].width = 25
    for i in range(1, len(projects) * 3 + 1):
        col_letter = chr(ord('A') + i)
        ws.column_dimensions[col_letter].width = 15
    
    # 第1行：项目名称
    ws['A1'] = "项目名称"
    for idx, project in enumerate(projects):
        col_idx = idx * 3 + 2
        cell = ws.cell(row=1, column=col_idx)
        cell.value = project["name"]
        cell.font = Font(bold=True)
    
    # 第2行：建筑面积
    ws['A2'] = "建筑面积(m²)"
    for idx, project in enumerate(projects):
        col_idx = idx * 3 + 2
        ws.cell(row=2, column=col_idx).value = project["area"]
    
    # 第3行：层数信息
    ws['A3'] = "层数"
    for idx, project in enumerate(projects):
        col_idx = idx * 3 + 2
        ws.cell(row=3, column=col_idx).value = project["floors"]
    
    # 第4行：开竣工时间
    ws['A4'] = "开竣工时间"
    for idx, project in enumerate(projects):
        col_idx = idx * 3 + 2
        ws.cell(row=4, column=col_idx).value = "2023-01-01至2024-01-01"
    
    # 成本数据
    cost_data = [
        ("1.0 土石方工程", [1000000, 800000, 1200000]),
        ("1.1 土方开挖", [600000, 500000, 700000]),
        ("1.2 土方回填", [400000, 300000, 500000]),
        ("2.0 地基与基础工程", [5000000, 4000000, 6000000]),
        ("2.1 桩基础工程", [3000000, 2500000, 3500000]),
        ("2.2 地下室工程", [2000000, 1500000, 2500000]),
        ("3.0 主体结构工程", [15000000, 12000000, 18000000]),
        ("4.0 建筑装饰装修工程", [8000000, 6500000, 9500000]),
        ("5.0 建筑屋面工程", [1500000, 1200000, 1800000]),
        ("6.0 建筑给排水工程", [2500000, 2000000, 3000000]),
        ("7.0 建筑电气工程", [3500000, 3000000, 4000000]),
        ("8.0 智能建筑工程", [2000000, 1800000, 2200000]),
        ("9.0 通风与空调工程", [4000000, 3500000, 4500000]),
        ("10.0 建筑节能工程", [1000000, 900000, 1100000]),
        ("11.0 电梯工程", [2500000, 2200000, 2800000]),
        ("12.0 室外工程", [1500000, 1300000, 1700000]),
        ("13.0 其他费用", [2000000, 1800000, 2200000]),
        ("14.0 项目总开发成本", [49000000, 41500000, 57300000]),
    ]
    
    # 填充数据
    current_row = 5
    for section_name, values in cost_data:
        ws.cell(row=current_row, column=1).value = section_name
        
        for idx, value in enumerate(values[:len(projects)]):
            col_idx = idx * 3 + 2
            area = projects[idx]["area"]
            
            # 合价
            ws.cell(row=current_row, column=col_idx).value = value
            ws.cell(row=current_row, column=col_idx).number_format = '#,##0.00'
            
            # 单价
            unit_price = value / area if area > 0 else 0
            ws.cell(row=current_row, column=col_idx + 1).value = unit_price
            ws.cell(row=current_row, column=col_idx + 1).number_format = '#,##0.00'
        
        current_row += 1
    
    wb.save(output_path)
    print(f"✅ 测试Excel文件已创建: {output_path}")


if __name__ == "__main__":
    # 创建测试数据目录
    fixtures_dir = Path(__file__).parent.parent.parent / "tests" / "unit" / "excel_parser" / "fixtures"
    fixtures_dir.mkdir(parents=True, exist_ok=True)
    
    print("🎯 开始创建测试Excel文件...")
    
    # 创建测试文件
    create_test_excel(str(fixtures_dir / "sample_3_projects.xlsx"), 3)
    create_test_excel(str(fixtures_dir / "sample_1_project.xlsx"), 1)
    
    print("🎉 测试Excel文件创建完成！")
    print(f"📁 文件位置: {fixtures_dir}")
